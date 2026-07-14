import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_tusk_amr_bom():
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Styling Palettes
    font_family = "Segoe UI"
    
    # Colors
    header_fill_color = "2C3E50"      # Dark Slate Blue/Gray
    accent_fill_color = "4A90E2"      # Steel Blue
    zebra_fill_color = "F8F9FA"       # Light gray
    calc_header_fill = "D9E1F2"       # Muted Blue-Gray for calculation sections
    calc_accent_fill = "E2EFDA"       # Muted Green for calculation outputs
    text_white = "FFFFFF"
    
    # Fonts
    title_font = Font(name=font_family, size=16, bold=True, color=text_white)
    subtitle_font = Font(name=font_family, size=11, italic=True, color="E0E0E0")
    header_font = Font(name=font_family, size=11, bold=True, color=text_white)
    section_font = Font(name=font_family, size=12, bold=True, color="2C3E50")
    bold_font = Font(name=font_family, size=10, bold=True)
    regular_font = Font(name=font_family, size=10)
    italic_font = Font(name=font_family, size=10, italic=True)
    
    # Fills
    header_fill = PatternFill(start_color=header_fill_color, end_color=header_fill_color, fill_type="solid")
    accent_fill = PatternFill(start_color=accent_fill_color, end_color=accent_fill_color, fill_type="solid")
    zebra_fill = PatternFill(start_color=zebra_fill_color, end_color=zebra_fill_color, fill_type="solid")
    calc_header_pattern = PatternFill(start_color=calc_header_fill, end_color=calc_header_fill, fill_type="solid")
    calc_accent_pattern = PatternFill(start_color=calc_accent_fill, end_color=calc_accent_fill, fill_type="solid")
    
    # Borders
    thin_side = Side(border_style="thin", color="D3D3D3")
    thick_bottom = Side(border_style="medium", color="2C3E50")
    double_bottom = Side(border_style="double", color="2C3E50")
    
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    header_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thick_bottom)
    total_border = Border(top=thin_side, bottom=double_bottom)
    
    # Alignments
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")
    wrap_left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ==========================================
    # SHEET 1: PROJECT SUMMARY
    # ==========================================
    ws_summary = wb.create_sheet(title="Project Summary")
    ws_summary.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws_summary.merge_cells("A1:D2")
    title_cell = ws_summary["A1"]
    title_cell.value = "Target Model: TUSK E10T (Telescopic Pallet APR)"
    title_cell.font = title_font
    title_cell.fill = header_fill
    title_cell.alignment = center_align
    
    ws_summary.row_dimensions[1].height = 25
    ws_summary.row_dimensions[2].height = 20
    
    # Specifications Table
    specs = [
        ("", ""),
        ("Design Parameter", "Specification Value"),
        ("Base Robot Model Reference", "TUSK E10T (Telescopic extending forks & folding wheels)"),
        ("Target Pallet Compatibility", "EUR-2 (1200x1000mm) & EUR-1 (1200x800mm) including double-sided pallets"),
        ("Maximum Payload Capacity", "1000 kg (1.0 Ton)"),
        ("Robot Self-Weight", "320 kg (including telescopic rails, carriage, and battery)"),
        ("Kinematic Layout", "Differential Drive (2x Active mid-wheels, 4x Corner spring-suspended casters)"),
        ("Maximum Speed (Loaded / Unloaded)", "1.5 m/s (Loaded) / 2.0 m/s (Unloaded)"),
        ("Lifting Height Range", "75 mm (Lowered) to 175 mm (Fully Raised)"),
        ("Lifting Mechanism", "Central vertical lifting carriage driven by a single motor and ball screw"),
        ("Fork Extension Mechanism", "Synchronized telescopic guide rails driven by a single motor via a splined shaft"),
        ("Fork Wheel Mechanism", "Passive retractable load rollers (rods pull rollers up when forks retract, and push them down during lift)"),
        ("Fork Width Adjustment Mechanism", "Transverse Lead Screw for sliding forks on carriage (Adjustable span 550mm - 900mm)"),
        ("Power Source", "48V 60Ah Lithium Iron Phosphate (LiFePO4) Battery with Smart BMS"),
        ("Navigation System", "Hybrid SLAM (2D LiDAR + 3D Depth Camera) & DM QR-code navigation"),
        ("Control Hardware", "ROS 2 Navigation PC (Intel Core i5) + STM32 Microcontroller (micro-ROS)"),
        ("Aisle Requirements", "Minimum Travel Aisle: 1.4m | Minimum Pick Aisle: 2.0m")
    ]
    
    for row_idx, (param, val) in enumerate(specs, start=3):
        ws_summary.cell(row=row_idx, column=1, value=param)
        ws_summary.cell(row=row_idx, column=2, value=val)
        
        # Formatting
        cell_p = ws_summary.cell(row=row_idx, column=1)
        cell_v = ws_summary.cell(row=row_idx, column=2)
        
        if row_idx == 4:
            cell_p.font = header_font
            cell_p.fill = accent_fill
            cell_p.alignment = center_align
            cell_v.font = header_font
            cell_v.fill = accent_fill
            cell_v.alignment = center_align
        elif row_idx > 4:
            cell_p.font = bold_font
            cell_p.fill = zebra_fill if row_idx % 2 == 0 else PatternFill(fill_type=None)
            cell_p.border = thin_border
            cell_p.alignment = left_align
            
            cell_v.font = regular_font
            cell_v.fill = zebra_fill if row_idx % 2 == 0 else PatternFill(fill_type=None)
            cell_v.border = thin_border
            cell_v.alignment = left_align
            
        ws_summary.row_dimensions[row_idx].height = 22

    # Adjust columns for Summary
    ws_summary.column_dimensions["A"].width = 38
    ws_summary.column_dimensions["B"].width = 65

    # ==========================================
    # SHEET 2: BILL OF MATERIALS (E10T SPECIFIC)
    # ==========================================
    ws_bom = wb.create_sheet(title="Bill of Materials")
    ws_bom.views.sheetView[0].showGridLines = True
    
    # Title row
    ws_bom.row_dimensions[1].height = 35
    ws_bom.merge_cells("A1:J1")
    bom_title = ws_bom["A1"]
    bom_title.value = "BOM: TUSK E10T Telescopic Pallet-Handling AMR (1-Ton)"
    bom_title.font = title_font
    bom_title.fill = header_fill
    bom_title.alignment = center_align
    
    headers = ["Item ID", "Category", "Component Name", "Description", "Specifications", "Supplier / Manufacturer", "Qty", "Unit Cost ($)", "Total Cost ($)", "Engineering Notes"]
    
    # Write headers
    ws_bom.row_dimensions[3].height = 26
    for col_idx, text in enumerate(headers, start=1):
        cell = ws_bom.cell(row=3, column=col_idx, value=text)
        cell.font = header_font
        cell.fill = accent_fill
        cell.alignment = center_align
        cell.border = header_border

    # BOM Items simplified to prevent confusion and represent a realistic design
    bom_data = [
        # Mechanical Structure
        ("Mechanical", "Chassis Base Frame", "Heavy-duty structural steel frame, laser-cut & welded plates", "10mm A36 Steel Plates, Custom Weldment", "Local CNC/Weld Shop", 1, 650.00, "Main frame holding battery, drive motors, and vertical guide rails"),
        ("Mechanical", "Vertical Lifting Carriage", "Slide weldment plate that raises/lowers the entire fork assembly", "8mm Steel Plates, Custom CNC Weldment", "Local CNC/Weld Shop", 1, 380.00, "Rides vertically on chassis rails; holds the fork width and extension assemblies"),
        ("Mechanical", "Suspended Left & Right Forks", "Low-profile sliding fork channels housing pull-rods", "Forks thickness 65mm, High-Strength Steel Q345", "Local CNC/Weld Shop", 2, 450.00, "Provides the support surface for lifting pallets"),
        ("Mechanical", "Fork Width Adjustment Slides", "Linear rails & carriages to slide forks transversely on carriage", "Hiwin HGR20 Rails (L=800mm) + HGW20CC Blocks", "Hiwin / local distributor", 2, 120.00, "Allows adjusting fork span for EUR-1 and EUR-2 pallets"),
        ("Mechanical", "Fork Extension Telescopic Slides", "Heavy-duty multi-stage telescopic linear rails for fork travel", "Rollon DSS43-1170 Telescopic Rails, Over-extension", "Rollon / local distributor", 2, 380.00, "Allows forks to extend 1400mm forward from the lift carriage"),
        ("Mechanical", "Fork Wheel Folding Linkages", "Mechanical swingarms and long pull-rods running inside forks", "Custom Q345 linkages + pinned joints", "Local CNC/Weld Shop", 2, 110.00, "Passively folds load rollers up when forks retract, and down when lifting"),
        ("Mechanical", "Central Lifting Ball Screw", "High-capacity vertical ball screw for carriage lift", "SFU2505 Ball Screw (L=400mm), C7 Accuracy", "TBI Motion / Misumi", 1, 95.00, "Single central screw to lift the entire carriage vertically"),
        ("Mechanical", "Vertical Guide Rails", "Heavy-duty linear profile rails for carriage vertical guide", "Hiwin HGR25 Rails (L=500mm) + HGW25CC Blocks", "Hiwin / local distributor", 2, 115.00, "Mounted vertically on chassis front face for lift stability"),
        ("Mechanical", "Heavy-Duty Drive Wheels", "Polyurethane tread wheels on cast iron core for drive traction", "Dia 200mm, Width 75mm, Keyway hub", "Blickle / local supplier", 2, 75.00, "Mid-drive wheels for differential steering layout"),
        ("Mechanical", "Heavy-Duty Caster Wheels", "Swivel casters with polyurethane tires for chassis corners", "Dia 100mm, load capacity 350kg each", "Blickle / Misumi", 4, 45.00, "Four corner support wheels with spring suspension"),
        ("Mechanical", "Fork Tip Load Rollers", "Low-profile tandem rollers inside fork tips, mounted on folding swingarms", "Dia 60mm, polyurethane, needle bearings", "Blickle / local supplier", 4, 30.00, "Sits under the front part of the forks, supports pallet weight"),
        ("Mechanical", "Bearing Block & Couplings", "Flanged support bearings for ball screws and flexible jaw couplings", "FK20/FF20 support blocks, D35 L50 Couplings", "Misumi / local supplier", 1, 70.00, "Connects central lift motor to the vertical ball screw"),
        
        # Actuation & Drive
        ("Actuation", "Differential Drive Motors", "Brushless DC servo motors with planetary gearboxes & electromagnetic brakes", "48V 750W, 3000RPM, 1:20 Planetary Gearbox, IP54", "Leadshine / ZYT Motor", 2, 380.00, "Provides 1.5 m/s speed and high starting torque"),
        ("Actuation", "Central Lifting BLDC Motor", "Brushless DC motor with planetary gearbox to drive vertical lift ball screw", "48V 1000W, 3000RPM, 1:15 Planetary Gearbox, Braked", "Leadshine / ZYT Motor", 1, 420.00, "Single motor to raise/lower the entire lifting carriage"),
        ("Actuation", "Fork Extension Geared Motor", "Geared motor driving a transverse splined shaft for synchronized fork travel", "24V 200W Geared DC Motor + Splined Drive Shaft", "Leadshine / local supplier", 1, 240.00, "Single motor to extend/retract both forks simultaneously"),
        ("Actuation", "Fork Width Adjustment Motor", "DC geared motor with integrated encoder for transverse lead screw drive", "24V 60W, 100RPM, planetary gear", "DFRobot / local supplier", 1, 95.00, "Adjusts the width slide of the forks dynamically"),
        
        # Electronics & Control
        ("Control", "Main Navigation Controller", "Industrial PC running Linux Ubuntu and ROS 2 Navigation stack", "Intel Core i5-1145G7, 16GB RAM, 256GB SSD, 9-36V DC", "Neousys / Advantech", 1, 750.00, "Runs SLAM, Nav2, path planning, and fleet interface"),
        ("Control", "Low-Level MCU Board", "Microcontroller board running micro-ROS for hardware interfaces", "Teensy 4.1 or STM32F407 Board", "STMicroelectronics / PJRC", 1, 45.00, "Handles encoder feedback, limit switches, and motor commands"),
        ("Control", "BLDC Motor Drivers", "High-performance dual-channel CAN-bus brushless motor controllers", "Roboteq SBL2360, 48V, Dual 60A per channel", "Roboteq", 2, 490.00, "1 driver for 2x drive motors; 1 driver for 1x lift motor & 1x extension motor"),
        ("Control", "Smart Battery Pack", "Lithium Iron Phosphate (LiFePO4) battery with RS485/CAN communication", "48V 60Ah, 150A continuous discharge, built-in Smart BMS", "Efest / RELiON / Custom", 1, 1200.00, "Provides 8+ hours of operating runtime"),
        ("Control", "Power Distribution Board", "Custom distribution board with circuit breakers, fuses, and relays", "Custom PCB, heavy-copper traces, 100A main contactor", "Local PCB shop", 1, 150.00, "Safely routes power to PC, drivers, and sensors"),
        ("Control", "Automatic Fast Charger", "Industrial smart battery charger with docking pads for auto-charging", "48V 15A output, contact pads for floor dock", "Delta Q / local supplier", 1, 350.00, "Interfaces with floor docking plate for automated charging"),
        
        # Sensors & Safety
        ("Sensors", "2D Navigation LiDARs", "Safety laser scanners for SLAM and 360-degree obstacle detection", "Sick TiM561 or Hokuyo UST-10LX, 10m range", "Sick / Hokuyo", 2, 980.00, "Mounted diagonally (front-right, rear-left) for full coverage"),
        ("Sensors", "3D Obstacle Avoidance Cameras", "Depth cameras for stereoscopic obstacle and height detection", "Intel RealSense D435i or Orbbec Astra Pro", "Intel / Orbbec", 2, 220.00, "Mounted front and rear for volumetric obstacle check"),
        ("Sensors", "Pallet Detection Sensors", "Diffuse reflection photoelectric sensors in forks", "Sick WL9 or Keyence PR-M, miniature photoeyes", "Sick / Keyence", 4, 85.00, "Detects pallet presence and docking alignment in fork channels"),
        ("Sensors", "High-Precision IMU", "Inertial Measurement Unit for SLAM odometry correction", "9-axis IMU, Bosch BNO055 or Xsens MTi-3", "Adafruit / Xsens", 1, 75.00, "Provides angular rate and linear acceleration data"),
        ("Sensors", "Inductive Limit Switches", "Proximity sensors for lift height limits and fork width calibration", "Omron E2E cylindrical proximity sensors", "Omron", 6, 25.00, "End-stop detection for mechanical linkages"),
        ("Sensors", "Safety Bumpers & E-Stops", "Pressure-sensitive contact edges and manual emergency stops", "TapeSwitch bumper strip + 2x red mushroom E-stops", "TapeSwitch / local supplier", 1, 180.00, "Physical safety overrides to stop motion immediately"),
        ("Sensors", "Warning Light Tower & Siren", "Audible and visual status indicator beacon", "RGB LED Tower, 85dB siren, 24V", "Banner Engineering / local", 1, 75.00, "Signals robot state (idle, moving, fault, charging)")
    ]

    current_row = 4
    for idx, item in enumerate(bom_data, start=1):
        ws_bom.cell(row=current_row, column=1, value=f"AMR-BOM-{idx:03d}").alignment = center_align
        ws_bom.cell(row=current_row, column=2, value=item[0]).alignment = center_align
        ws_bom.cell(row=current_row, column=3, value=item[1]).alignment = left_align
        ws_bom.cell(row=current_row, column=4, value=item[2]).alignment = wrap_left_align
        ws_bom.cell(row=current_row, column=5, value=item[3]).alignment = wrap_left_align
        ws_bom.cell(row=current_row, column=6, value=item[4]).alignment = left_align
        
        qty_cell = ws_bom.cell(row=current_row, column=7, value=item[5])
        qty_cell.alignment = center_align
        qty_cell.number_format = '#,##0'
        
        unit_cell = ws_bom.cell(row=current_row, column=8, value=item[6])
        unit_cell.alignment = right_align
        unit_cell.number_format = '$#,##0.00'
        
        # Formula for Total Cost: Qty * Unit Cost
        total_cell = ws_bom.cell(row=current_row, column=9, value=f"=G{current_row}*H{current_row}")
        total_cell.alignment = right_align
        total_cell.number_format = '$#,##0.00'
        total_cell.font = bold_font
        
        ws_bom.cell(row=current_row, column=10, value=item[7]).alignment = wrap_left_align
        
        # Apply fonts and borders
        for col in range(1, 11):
            c = ws_bom.cell(row=current_row, column=col)
            c.font = bold_font if col in [1, 9] else regular_font
            c.border = thin_border
            if current_row % 2 == 0:
                c.fill = zebra_fill
                
        ws_bom.row_dimensions[current_row].height = 36
        current_row += 1

    # Empty row
    current_row += 1
    
    # Grand Total Row
    ws_bom.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
    total_label = ws_bom.cell(row=current_row, column=1, value="GRAND TOTAL ESTIMATED HARDWARE COST")
    total_label.font = Font(name=font_family, size=11, bold=True, color="2C3E50")
    total_label.alignment = right_align
    
    total_val = ws_bom.cell(row=current_row, column=9, value=f"=SUM(I4:I{current_row-2})")
    total_val.font = Font(name=font_family, size=11, bold=True, color="2C3E50")
    total_val.alignment = right_align
    total_val.number_format = '$#,##0.00'
    total_val.border = total_border
    total_val.fill = calc_accent_pattern
    
    # Apply total border to the total label cell merged range
    for col in range(1, 9):
        ws_bom.cell(row=current_row, column=col).border = Border(top=thin_side, bottom=double_bottom)
        
    ws_bom.row_dimensions[current_row].height = 25

    # Auto-adjust column widths for BOM sheet
    column_widths = {
        "A": 14,  # Item ID
        "B": 13,  # Category
        "C": 28,  # Component Name
        "D": 36,  # Description
        "E": 34,  # Specifications
        "F": 22,  # Supplier
        "G": 8,   # Qty
        "H": 13,  # Unit Cost
        "I": 15,  # Total Cost
        "J": 42   # Engineering Notes
    }
    for col_letter, width in column_widths.items():
        ws_bom.column_dimensions[col_letter].width = width

    # ==========================================
    # SHEET 3: DESIGN CALCULATIONS
    # ==========================================
    ws_calc = wb.create_sheet(title="Kinematics & Calculations")
    ws_calc.views.sheetView[0].showGridLines = True
    
    ws_calc.row_dimensions[1].height = 30
    ws_calc.merge_cells("A1:E1")
    calc_title = ws_calc["A1"]
    calc_title.value = "E10T Sizing & Engineering Calculations"
    calc_title.font = title_font
    calc_title.fill = header_fill
    calc_title.alignment = center_align

    # Section 1: Central Vertical Lift Sizing
    ws_calc.cell(row=3, column=1, value="1. Central Vertical Lift Carriage Sizing").font = section_font
    
    calc_headers = ["Parameter / Step", "Symbol / Formula", "Value", "Unit", "Explanation"]
    ws_calc.row_dimensions[4].height = 24
    for col_idx, h in enumerate(calc_headers, start=1):
        c = ws_calc.cell(row=4, column=col_idx, value=h)
        c.font = header_font
        c.fill = accent_fill
        c.alignment = center_align
        c.border = header_border

    scissor_data = [
        ("Total Load to Lift", "W = (Payload + Lift Carriage)", 1080, "kg", "1000kg max payload + 80kg carriage and forks weight"),
        ("Total Lifting Weight Force", "F_g = W * 9.81", "=C5*9.81", "N", "Total gravitational force of load"),
        ("Required Lifting Force", "F_lift = F_g", "=C6", "N", "Vertical force to lift the entire carriage (direct vertical lift)"),
        ("Lifting Ball Screw Pitch", "p", 5, "mm", "Linear travel per rotation of the central vertical ball screw"),
        ("Lifting Ball Screw Lead", "L = p / 1000", "=C8/1000", "m", "Ball screw pitch in meters"),
        ("Ball Screw Mechanical Efficiency", "eta_screw", 0.90, "decimal", "Frictional efficiency of rolling balls in screw threads"),
        ("Screw Drive Torque Required", "T_screw = (F_lift * L) / (2 * pi * eta_screw)", "=(C7*C9)/(2*PI()*C10)", "Nm", "Torque required on the vertical ball screw shaft to lift load"),
        ("Lifting Gearbox Reduction Ratio", "R_gearbox", 15, "ratio", "Gear ratio of planetary gearbox on lift motor"),
        ("Gearbox Efficiency", "eta_gearbox", 0.95, "decimal", "Planetary gear stage efficiency"),
        ("Peak Lift Motor Torque Required", "T_motor = T_screw / (R_gearbox * eta_gearbox)", "=C11/(C12*C13)", "Nm", "Required output torque from the single central BLDC motor before gearing")
    ]

    r = 5
    for item in scissor_data:
        ws_calc.cell(row=r, column=1, value=item[0]).font = bold_font
        ws_calc.cell(row=r, column=2, value=item[1]).font = italic_font
        ws_calc.cell(row=r, column=4, value=item[3]).font = regular_font
        ws_calc.cell(row=r, column=5, value=item[4]).font = regular_font
        
        val_cell = ws_calc.cell(row=r, column=3, value=item[2])
        val_cell.font = bold_font
        val_cell.alignment = right_align
        
        # Number formats
        if r in [5]:
            val_cell.number_format = '#,##0'
        elif r in [6, 7]:
            val_cell.number_format = '#,##0.0'
        elif r == 8:
            val_cell.number_format = '0'
        elif r in [9, 10, 13]:
            val_cell.number_format = '0.00'
        elif r == 11:
            val_cell.number_format = '0.00'
        elif r == 14:
            val_cell.number_format = '0.00'
            val_cell.fill = calc_accent_pattern
            
        for col in range(1, 6):
            ws_calc.cell(row=r, column=col).border = thin_border
            if r % 2 == 0:
                ws_calc.cell(row=r, column=col).fill = zebra_fill
        r += 1

    # Section 2: Drive Traction Calculations
    r += 1
    ws_calc.cell(row=r, column=1, value="2. Differential Drive Traction Sizing").font = section_font
    r += 1
    
    # Drive headers
    ws_calc.row_dimensions[r].height = 24
    for col_idx, h in enumerate(calc_headers, start=1):
        c = ws_calc.cell(row=r, column=col_idx, value=h)
        c.font = header_font
        c.fill = accent_fill
        c.alignment = center_align
        c.border = header_border
    r += 1
    
    drive_start_row = r
    drive_data = [
        ("Total Gross Mass (AMR + Payload)", "M = M_amr + M_payload", 1320, "kg", "Robot mass 320kg (E10T weight) + 1000kg payload"),
        ("Maximum Travel Speed", "v_max", 1.5, "m/s", "Design speed of AMR at full capacity"),
        ("Acceleration Time to Max Speed", "t_acc", 2.0, "s", "Time limit to reach full speed"),
        ("Required Linear Acceleration", "a = v_max / t_acc", f"=C{drive_start_row+1}/C{drive_start_row+2}", "m/s^2", "Target acceleration rate"),
        ("Acceleration Force Required", "F_acc = M * a", f"=C{drive_start_row}*C{drive_start_row+3}", "N", "Force needed to accelerate mass"),
        ("Rolling Resistance Coefficient", "f_r", 0.02, "decimal", "Polyurethane wheels on smooth, hard industrial concrete floor"),
        ("Rolling Resistance Force", "F_roll = M * g * f_r", f"=C{drive_start_row}*9.81*C{drive_start_row+5}", "N", "Force to overcome rolling friction"),
        ("Total Tractive Force", "F_total = F_acc + F_roll", f"=C{drive_start_row+4}+C{drive_start_row+6}", "N", "Sum of acceleration and rolling friction forces"),
        ("Number of Driving Motors/Wheels", "N_drive", 2, "qty", "Differential drive layout has two independent driven wheels"),
        ("Tractive Force per Drive Wheel", "F_wheel = F_total / N_drive", f"=C{drive_start_row+7}/C{drive_start_row+8}", "N", "Tractive force share per driving wheel"),
        ("Drive Wheel Radius", "R_wheel", 0.10, "m", "100mm wheel radius (200mm diameter drive wheel)"),
        ("Peak Wheel Torque Required", "T_wheel = F_wheel * R_wheel", f"=C{drive_start_row+9}*C{drive_start_row+10}", "Nm", "Required torque at the wheel axle"),
        ("Drive Gearbox Reduction Ratio", "R_drive_gearbox", 20, "ratio", "Planetary gearbox reduction ratio"),
        ("Drive Gearbox Mechanical Efficiency", "eta_drive_box", 0.95, "decimal", "Efficiency of planetary stages"),
        ("Peak Drive Motor Torque Required", "T_drive_motor = T_wheel / (R_drive_gearbox * eta_drive_box)", f"=C{drive_start_row+11}/(C{drive_start_row+12}*C{drive_start_row+13})", "Nm", "Required motor electromagnetic output torque"),
        ("Rated Motor Rotation Speed", "omega_motor = (v_max / R_wheel) * R_drive_gearbox * 60 / (2*pi)", f"=(C{drive_start_row+1}/C{drive_start_row+10})*C{drive_start_row+12}*60/(2*PI())", "RPM", "Rotational speed of motor shaft at max linear velocity"),
        ("Mechanical Power Output per Motor", "P_motor = T_drive_motor * (omega_motor * 2*pi / 60)", f"=C{drive_start_row+14}*C{drive_start_row+15}*2*PI()/60", "W", "Continuous mechanical wattage required per motor")
    ]

    for item in drive_data:
        ws_calc.cell(row=r, column=1, value=item[0]).font = bold_font
        ws_calc.cell(row=r, column=2, value=item[1]).font = italic_font
        ws_calc.cell(row=r, column=4, value=item[3]).font = regular_font
        ws_calc.cell(row=r, column=5, value=item[4]).font = regular_font
        
        val_cell = ws_calc.cell(row=r, column=3, value=item[2])
        val_cell.font = bold_font
        val_cell.alignment = right_align
        
        # Formatting specific columns
        if r in [drive_start_row, drive_start_row+8]:
            val_cell.number_format = '#,##0'
        elif r in [drive_start_row+1, drive_start_row+2, drive_start_row+10]:
            val_cell.number_format = '0.0'
        elif r in [drive_start_row+3, drive_start_row+5, drive_start_row+13]:
            val_cell.number_format = '0.00'
        elif r in [drive_start_row+4, drive_start_row+6, drive_start_row+7, drive_start_row+9]:
            val_cell.number_format = '#,##0.0'
        elif r in [drive_start_row+11, drive_start_row+14]:
            val_cell.number_format = '0.00'
        elif r == drive_start_row+15:
            val_cell.number_format = '#,##0'
        elif r == drive_start_row+16:
            val_cell.number_format = '#,##0'
            val_cell.fill = calc_accent_pattern
            
        for col in range(1, 6):
            ws_calc.cell(row=r, column=col).border = thin_border
            if r % 2 == 0:
                ws_calc.cell(row=r, column=col).fill = zebra_fill
        r += 1

    # Adjust calculations column width
    ws_calc.column_dimensions["A"].width = 38
    ws_calc.column_dimensions["B"].width = 46
    ws_calc.column_dimensions["C"].width = 16
    ws_calc.column_dimensions["D"].width = 12
    ws_calc.column_dimensions["E"].width = 54

    # ==========================================
    # SHEET 4: WIRING ARCHITECTURE
    # ==========================================
    ws_wire = wb.create_sheet(title="Wiring Architecture")
    ws_wire.views.sheetView[0].showGridLines = True
    
    ws_wire.row_dimensions[1].height = 30
    ws_wire.merge_cells("A1:F1")
    wire_title = ws_wire["A1"]
    wire_title.value = "E10T Electrical Connections & Signal Routing"
    wire_title.font = title_font
    wire_title.fill = header_fill
    wire_title.alignment = center_align
    
    wire_headers = ["Connection ID", "Source Device", "Destination Device", "Cable Type", "Interface Protocol", "Function / Description"]
    ws_wire.row_dimensions[3].height = 24
    for col_idx, h in enumerate(wire_headers, start=1):
        c = ws_wire.cell(row=3, column=col_idx, value=h)
        c.font = header_font
        c.fill = accent_fill
        c.alignment = center_align
        c.border = header_border
        
    wire_data = [
        ("PWR-01", "Smart Battery Pack (48V)", "Main Circuit Breaker / Switch", "AWG 4 Silicon Wire", "DC Power", "Main power input connection"),
        ("PWR-02", "Circuit Breaker", "Power Distribution Board", "AWG 6 Silicon Wire", "DC Power", "Fused main DC distribution bus"),
        ("PWR-03", "Power Distribution Board", "Drive Motor Drivers (Roboteq)", "AWG 10 Silicon Wire", "DC Power", "High current feed for drive motors (fused)"),
        ("PWR-04", "Power Distribution Board", "Lifting Motor Drivers (Roboteq)", "AWG 10 Silicon Wire", "DC Power", "High current feed for lifting motors (fused)"),
        ("PWR-05", "Power Distribution Board", "Voltage Regulator (48V to 12V/24V)", "AWG 16 wire", "DC Power", "Power feed to step-down converters"),
        ("PWR-06", "Voltage Regulator (24V)", "Industrial PC (Nav Controller)", "AWG 18 wire", "24V DC Power", "Regulated logic power for IPC"),
        ("PWR-07", "Voltage Regulator (24V)", "Sensors & Low-level Controllers", "AWG 20 wire", "24V/12V Power", "Power for LiDARs, cameras, and MCU"),
        
        ("SIG-01", "2D Navigation LiDARs (2x)", "Industrial PC", "Cat6 Shielded RJ45", "Ethernet (TCP/IP)", "Point cloud data transmission for SLAM"),
        ("SIG-02", "3D Obstacle Cameras (2x)", "Industrial PC", "USB 3.0 Type-C Cable", "USB 3.0", "RGB-D frames for stereoscopic obstacle checks"),
        ("SIG-03", "IMU Sensor", "Low-Level MCU Board", "4-pin Ribbon Cable", "I2C / UART", "Raw accelerometer/gyroscope measurements"),
        ("SIG-04", "Wheel Encoders (2x)", "Low-Level MCU Board", "Shielded 6-wire Cable", "Quadrature (A/B)", "Incremental drive wheel feedback for odometry"),
        ("SIG-05", "Lift Linear Encoder", "Low-Level MCU Board", "Shielded 6-wire Cable", "SSI / SSI-SPI", "Absolute vertical position of the lifting carriage"),
        ("SIG-06", "Limit & Proximity Switches (6x)", "Low-Level MCU Board", "3-wire Sensor Cable", "Digital I/O (24V)", "End-of-travel safety flags for lift and slide"),
        ("SIG-07", "Pallet Photoelectric Sensors (4x)", "Low-Level MCU Board", "3-wire Sensor Cable", "Digital I/O (24V)", "Detects pallet legs and correct insertion"),
        
        ("COM-01", "Industrial PC", "Low-Level MCU Board", "Shielded USB Cable", "micro-ROS Serial (USB)", "Command velocities (cmd_vel) and odometry feedback"),
        ("COM-02", "Low-Level MCU Board", "Drive Motor Drivers", "Shielded twisted pair", "CAN-bus", "Position/velocity control frames for wheels"),
        ("COM-03", "Low-Level MCU Board", "Lifting Motor Driver", "Shielded twisted pair", "CAN-bus", "Vertical lift motion control"),
        ("COM-04", "Smart Battery BMS", "Industrial PC", "RS485 to USB converter", "Modbus RTU / CAN", "State of Charge (SoC), voltage, and diagnostic logs")
    ]

    current_row = 4
    for idx, item in enumerate(wire_data, start=1):
        ws_wire.cell(row=current_row, column=1, value=item[0]).alignment = center_align
        ws_wire.cell(row=current_row, column=2, value=item[1]).alignment = left_align
        ws_wire.cell(row=current_row, column=3, value=item[2]).alignment = left_align
        ws_wire.cell(row=current_row, column=4, value=item[3]).alignment = left_align
        ws_wire.cell(row=current_row, column=5, value=item[4]).alignment = center_align
        ws_wire.cell(row=current_row, column=6, value=item[5]).alignment = wrap_left_align
        
        for col in range(1, 7):
            c = ws_wire.cell(row=current_row, column=col)
            c.font = bold_font if col == 1 else regular_font
            c.border = thin_border
            if current_row % 2 == 0:
                c.fill = zebra_fill
                
        ws_wire.row_dimensions[current_row].height = 24
        current_row += 1

    # Adjust wiring columns width
    ws_wire.column_dimensions["A"].width = 16
    ws_wire.column_dimensions["B"].width = 28
    ws_wire.column_dimensions["C"].width = 28
    ws_wire.column_dimensions["D"].width = 22
    ws_wire.column_dimensions["E"].width = 22
    ws_wire.column_dimensions["F"].width = 54

    # Save Workbook
    filepath = "/home/charan/.gemini/antigravity-ide/scratch/tusk-amr-analysis/tusk_amr_bom.xlsx"
    wb.save(filepath)
    print(f"Workbook successfully saved to {filepath}")

if __name__ == "__main__":
    create_tusk_amr_bom()
