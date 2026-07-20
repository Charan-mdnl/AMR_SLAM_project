import openpyxl
import re
import os

def markdown_to_html(md_text):
    # Basic markdown parser to convert design_report.md to styled HTML
    html_content = ""
    lines = md_text.split('\n')
    in_list = False
    in_table = False
    in_code = False
    table_headers = []
    
    for line in lines:
        stripped = line.strip()
        
        # Code block
        if stripped.startswith("```"):
            if in_code:
                html_content += "</pre></code></div>"
                in_code = False
            else:
                lang = stripped[3:].strip()
                html_content += f"<div class='code-block'><code class='language-{lang}'><pre>"
                in_code = True
            continue
            
        if in_code:
            html_content += line + "\n"
            continue
            
        # Lists
        if stripped.startswith("* ") or stripped.startswith("- "):
            if not in_list:
                html_content += "<ul>"
                in_list = True
            content = stripped[2:]
            # Bold/Italic inside list
            content = re.sub(r'\*\*(.*?)\*\*', r'<strong>\1</strong>', content)
            content = re.sub(r'\*(.*?)\*', r'<em>\1</em>', content)
            content = re.sub(r'\$(.*?)\$', r'<span class="math">\1</span>', content)
            html_content += f"<li>{content}</li>"
            continue
        elif in_list and not (stripped.startswith("* ") or stripped.startswith("- ") or stripped == ""):
            html_content += "</ul>"
            in_list = False

        # Tables
        if stripped.startswith("|"):
            if not in_table:
                html_content += "<table class='report-table'>"
                in_table = True
            
            # Parse cells
            cells = [c.strip() for c in stripped.split("|")[1:-1]]
            if len(cells) > 0:
                if stripped.replace(" ", "").startswith("|:---") or stripped.replace(" ", "").startswith("|---"):
                    # Table separator, skip
                    continue
                
                # Check if it's header row (first row of table)
                if not table_headers:
                    table_headers = cells
                    html_content += "<thead><tr>"
                    for cell in cells:
                        html_content += f"<th>{cell}</th>"
                    html_content += "</tr></thead><tbody>"
                else:
                    html_content += "<tr>"
                    for cell in cells:
                        # Parse math/bold in cells
                        cell_fmt = re.sub(r'\*\*(.*?)\*\*', r'<strong>\1</strong>', cell)
                        cell_fmt = re.sub(r'\$(.*?)\$', r'<span class="math">\1</span>', cell_fmt)
                        html_content += f"<td>{cell_fmt}</td>"
                    html_content += "</tr>"
            continue
        elif in_table and not stripped.startswith("|"):
            html_content += "</tbody></table>"
            in_table = False
            table_headers = []

        # Headings
        if stripped.startswith("# "):
            html_content += f"<h1>{stripped[2:]}</h1>"
        elif stripped.startswith("## "):
            html_content += f"<h2>{stripped[3:]}</h2>"
        elif stripped.startswith("### "):
            html_content += f"<h3>{stripped[4:]}</h3>"
        elif stripped.startswith("#### "):
            html_content += f"<h4>{stripped[5:]}</h4>"
            
        # Divider
        elif stripped == "---":
            html_content += "<hr>"
            
        # Alert Blocks
        elif stripped.startswith("> [!IMPORTANT]"):
            html_content += "<div class='alert alert-important'><strong>Important:</strong> "
        elif stripped.startswith("> [!NOTE]"):
            html_content += "<div class='alert alert-note'><strong>Note:</strong> "
        elif stripped.startswith("> ") and (html_content.endswith("important'><strong>Important:</strong> ") or html_content.endswith("note'><strong>Note:</strong> ") or html_content.endswith("<br>")):
            content = stripped[2:]
            html_content += f"{content}<br>"
        elif stripped == "" and (html_content.endswith("<br>") or "alert" in html_content[-100:]):
            if html_content.endswith("<br>"):
                html_content = html_content[:-4]
            html_content += "</div>"
            
        # Markdown Image
        elif stripped.startswith("![") and stripped.endswith(")"):
            match = re.match(r'!\[(.*?)\]\((.*?)\)', stripped)
            if match:
                alt = match.group(1)
                src = match.group(2)
                html_content += f"<div class='img-container'><img src='{src}' alt='{alt}' class='report-image' /><div class='img-caption'>{alt}</div></div>"
            
        # Paragraphs
        elif stripped != "":
            # Bold/Italic and inline math
            line_fmt = re.sub(r'\*\*(.*?)\*\*', r'<strong>\1</strong>', stripped)
            line_fmt = re.sub(r'\*(.*?)\*', r'<em>\1</em>', line_fmt)
            line_fmt = re.sub(r'\$(.*?)\$', r'<span class="math">\1</span>', line_fmt)
            html_content += f"<p>{line_fmt}</p>"
            
    return html_content

def build_report_html(report_path, html_out, title):
    if not os.path.exists(report_path):
        print(f"Error: {report_path} not found")
        return
        
    with open(report_path, "r", encoding="utf-8") as f:
        md_text = f.read()
        
    body_content = markdown_to_html(md_text)
    
    html_page = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{title}</title>
    <link href="https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;500;700&family=JetBrains+Mono:wght@400;700&display=swap" rel="stylesheet">
    <style>
        body {{
            font-family: 'Outfit', sans-serif;
            background-color: #0f172a;
            color: #e2e8f0;
            line-height: 1.6;
            margin: 0;
            padding: 40px 20px;
        }}
        .container {{
            max-width: 900px;
            margin: 0 auto;
            background: rgba(30, 41, 59, 0.7);
            padding: 40px;
            border-radius: 16px;
            box-shadow: 0 8px 32px 0 rgba(0, 0, 0, 0.37);
            backdrop-filter: blur(8px);
            border: 1px solid rgba(255, 255, 255, 0.1);
        }}
        h1 {{
            font-size: 2.2rem;
            color: #38bdf8;
            border-bottom: 2px solid rgba(56, 189, 248, 0.3);
            padding-bottom: 10px;
            margin-top: 0;
        }}
        h2 {{
            font-size: 1.6rem;
            color: #38bdf8;
            margin-top: 30px;
            border-bottom: 1px solid rgba(255, 255, 255, 0.1);
            padding-bottom: 5px;
        }}
        h3 {{
            font-size: 1.25rem;
            color: #f1f5f9;
            margin-top: 20px;
        }}
        p {{
            color: #cbd5e1;
            margin-bottom: 20px;
        }}
        ul {{
            margin-bottom: 20px;
            padding-left: 20px;
        }}
        li {{
            margin-bottom: 8px;
            color: #cbd5e1;
        }}
        .math {{
            font-family: 'JetBrains Mono', monospace;
            background: rgba(15, 23, 42, 0.6);
            padding: 2px 6px;
            border-radius: 4px;
            color: #f472b6;
            font-size: 0.9em;
        }}
        hr {{
            border: 0;
            height: 1px;
            background: rgba(255, 255, 255, 0.1);
            margin: 30px 0;
        }}
        .alert {{
            padding: 15px;
            border-radius: 8px;
            margin: 20px 0;
            border-left: 4px solid;
        }}
        .alert-important {{
            background: rgba(239, 68, 68, 0.15);
            border-color: #ef4444;
            color: #fca5a5;
        }}
        .alert-note {{
            background: rgba(59, 130, 246, 0.15);
            border-color: #3b82f6;
            color: #93c5fd;
        }}
        .report-table {{
            width: 100%;
            border-collapse: collapse;
            margin: 25px 0;
            font-size: 0.95rem;
        }}
        .report-table th, .report-table td {{
            padding: 12px 15px;
            text-align: left;
            border-bottom: 1px solid rgba(255, 255, 255, 0.1);
        }}
        .report-table th {{
            background-color: rgba(15, 23, 42, 0.6);
            color: #38bdf8;
            font-weight: 500;
        }}
        .report-table tr:hover {{
            background-color: rgba(255, 255, 255, 0.02);
        }}
        .code-block {{
            background: #090d16;
            padding: 20px;
            border-radius: 8px;
            overflow-x: auto;
            border: 1px solid rgba(255, 255, 255, 0.05);
            margin: 20px 0;
        }}
        pre, code {{
            font-family: 'JetBrains Mono', monospace;
            margin: 0;
            color: #a7f3d0;
        }}
        .img-container {{
            text-align: center;
            margin: 30px 0;
            background: rgba(15, 23, 42, 0.4);
            padding: 20px;
            border-radius: 12px;
            border: 1px solid rgba(255, 255, 255, 0.05);
            page-break-inside: avoid;
        }}
        .report-image {{
            max-width: 100%;
            max-height: 550px;
            height: auto;
            border-radius: 8px;
            border: 1px solid rgba(255, 255, 255, 0.1);
        }}
        .img-caption {{
            font-size: 0.9rem;
            color: #94a3b8;
            margin-top: 12px;
            font-style: italic;
            font-weight: 500;
        }}
        .btn-print {{
            display: inline-block;
            background: #0284c7;
            color: white;
            padding: 10px 20px;
            border-radius: 6px;
            text-decoration: none;
            font-weight: 500;
            margin-bottom: 25px;
            transition: background 0.2s;
            cursor: pointer;
            border: none;
            font-family: inherit;
        }}
        .btn-print:hover {{
            background: #0369a1;
        }}
        @media print {{
            @page {{
                size: A4 portrait;
                margin: 20mm 15mm 20mm 15mm;
            }}
            body {{
                background-color: white;
                color: black;
                padding: 0;
                margin: 0;
            }}
            .container {{
                background: none;
                box-shadow: none;
                padding: 0;
                border: none;
                max-width: 100%;
            }}
            .btn-print {{
                display: none;
            }}
            h1, h2, h3, .report-table th {{
                color: black;
                page-break-after: avoid;
            }}
            p, li, td {{
                color: #222;
            }}
            tr, li, .code-block, .img-container {{
                page-break-inside: avoid;
            }}
            .math {{
                background: #f1f5f9;
                color: #b91c1c;
            }}
            .code-block {{
                background: #fafafa;
                border: 1px solid #ddd;
            }}
            code {{
                color: #111827;
            }}
            .img-container {{
                background: none;
                border: none;
                padding: 10px 0;
            }}
            .report-image {{
                border: 1px solid #ccc;
            }}
            .img-caption {{
                color: #444;
            }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <button class="btn-print" onclick="window.print()">Print / Export as PDF</button>
        {body_content}
    </div>
</body>
</html>
"""
    with open(html_out, "w", encoding="utf-8") as f:
        f.write(html_page)
    print(f"Report HTML generated successfully: {html_out}")

def excel_to_html_tabs():
    excel_path = "/home/charan/.gemini/antigravity-ide/scratch/tusk-amr-analysis/tusk_amr_bom.xlsx"
    html_out = "/home/charan/.gemini/antigravity-ide/scratch/tusk-amr-analysis/tusk_amr_bom.html"
    
    if not os.path.exists(excel_path):
        print(f"Error: {excel_path} not found")
        return
        
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    tab_headers_html = ""
    tab_contents_html = ""
    
    for idx, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        active_class = "active" if idx == 0 else ""
        
        # Build Tab Button
        tab_headers_html += f"""
        <button class="tab-link {active_class}" onclick="openTab(event, 'tab_{idx}')">{sheet_name}</button>
        """
        
        # Build Tab Content (HTML Table)
        table_rows = ""
        
        # Find dimension
        max_r = ws.max_row
        max_c = ws.max_column
        
        # Row heights/col widths for rendering adjustment
        for r in range(1, max_r + 1):
            row_html = "<tr>"
            
            # Row height
            rh = ws.row_dimensions[r].height
            rh_style = f"style='height:{rh}px;'" if rh else ""
            
            # Check if this row has merged cells
            is_merged_header = False
            merged_val = None
            merged_span = 1
            
            # We will handle merged cells simply: if a cell is merged, we check if it's the start
            for c in range(1, max_c + 1):
                cell = ws.cell(row=r, column=c)
                val = cell.value if cell.value is not None else ""
                
                # Check for formula indicators or format as currency/numbers
                val_str = str(val)
                
                # Determine align & styling based on cell properties
                align = "left"
                if cell.alignment and cell.alignment.horizontal:
                    align = cell.alignment.horizontal
                
                bold_style = "font-weight: bold;" if cell.font and cell.font.bold else ""
                italic_style = "font-style: italic;" if cell.font and cell.font.italic else ""
                
                # Background fill conversion
                bg_style = ""
                if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
                    rgb = cell.fill.start_color.rgb
                    if isinstance(rgb, str) and len(rgb) in (6, 8) and all(c in '0123456789ABCDEFabcdef' for c in rgb):
                        if len(rgb) == 8:
                            rgb = rgb[2:] # Remove alpha
                        if rgb != "00000000" and rgb != "000000":
                            bg_style = f"background-color: #{rgb};"
                
                # Border styling
                border_style = "border: 1px solid rgba(255,255,255,0.1);"
                
                # Color text
                color_style = ""
                if cell.font and cell.font.color and cell.font.color.rgb:
                    color_rgb = cell.font.color.rgb
                    if isinstance(color_rgb, str) and len(color_rgb) in (6, 8) and all(c in '0123456789ABCDEFabcdef' for c in color_rgb):
                        if len(color_rgb) == 8:
                            color_rgb = color_rgb[2:]
                        if color_rgb != "00000000" and color_rgb != "FFFFFFFF" and color_rgb != "000000":
                            color_style = f"color: #{color_rgb};"
                        elif color_rgb == "FFFFFFFF":
                            color_style = "color: #ffffff;"
                
                # Apply number formatting in python representation
                if isinstance(val, float) or isinstance(val, int):
                    if cell.number_format and '$' in cell.number_format:
                        val_str = f"${val:,.2f}"
                    elif cell.number_format and '%' in cell.number_format:
                        val_str = f"{val*100:.1f}%"
                    elif isinstance(val, float):
                        val_str = f"{val:.2f}"
                
                # Detect merged cell ranges in sheet
                is_merged = False
                for merged_range in ws.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        is_merged = True
                        # If this cell is the top-left of the range, we render it with colspan/rowspan
                        if cell.coordinate == merged_range.start_cell.coordinate:
                            colspan = merged_range.size['columns']
                            rowspan = merged_range.size['rows']
                            row_html += f"<td colspan='{colspan}' rowspan='{rowspan}' style='text-align:{align}; {bold_style} {italic_style} {bg_style} {color_style}' class='excel-cell'>{val_str}</td>"
                        break
                
                if not is_merged:
                    row_html += f"<td style='text-align:{align}; {bold_style} {italic_style} {bg_style} {color_style}' class='excel-cell'>{val_str}</td>"
                    
            row_html += "</tr>"
            table_rows += row_html
            
        tab_contents_html += f"""
        <div id="tab_{idx}" data-sheet-name="{sheet_name}" class="tab-content {active_class}">
            <div class="table-responsive">
                <table class="excel-table">
                    {table_rows}
                </table>
            </div>
        </div>
        """

    html_page = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>AMR Bill of Materials & Sizing Workbook</title>
    <link href="https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;500;700&family=JetBrains+Mono:wght@400;700&display=swap" rel="stylesheet">
    <style>
        body {{
            font-family: 'Outfit', sans-serif;
            background-color: #0f172a;
            color: #e2e8f0;
            margin: 0;
            padding: 30px;
        }}
        .header {{
            max-width: 1200px;
            margin: 0 auto 20px auto;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }}
        h1 {{
            font-size: 1.8rem;
            color: #38bdf8;
            margin: 0;
        }}
        .btn-print {{
            background: #0284c7;
            color: white;
            padding: 8px 16px;
            border-radius: 6px;
            text-decoration: none;
            font-weight: 500;
            cursor: pointer;
            border: none;
            font-family: inherit;
            transition: background 0.2s;
        }}
        .btn-print:hover {{
            background: #0369a1;
        }}
        .container {{
            max-width: 1200px;
            margin: 0 auto;
            background: rgba(30, 41, 59, 0.7);
            border-radius: 12px;
            box-shadow: 0 8px 32px 0 rgba(0, 0, 0, 0.37);
            backdrop-filter: blur(8px);
            border: 1px solid rgba(255, 255, 255, 0.1);
            overflow: hidden;
        }}
        .tabs {{
            display: flex;
            background: rgba(15, 23, 42, 0.8);
            border-bottom: 1px solid rgba(255, 255, 255, 0.1);
        }}
        .tab-link {{
            background: none;
            border: none;
            outline: none;
            padding: 14px 20px;
            color: #94a3b8;
            font-family: inherit;
            font-size: 1rem;
            font-weight: 500;
            cursor: pointer;
            transition: all 0.2s;
            border-bottom: 2px solid transparent;
        }}
        .tab-link:hover {{
            color: #e2e8f0;
            background: rgba(255, 255, 255, 0.02);
        }}
        .tab-link.active {{
            color: #38bdf8;
            border-bottom: 2px solid #38bdf8;
            background: rgba(255, 255, 255, 0.05);
        }}
        .tab-content {{
            display: none;
            padding: 20px;
        }}
        .tab-content.active {{
            display: block;
        }}
        .table-responsive {{
            overflow-x: auto;
            border-radius: 6px;
            border: 1px solid rgba(255, 255, 255, 0.05);
            background: #090d16;
        }}
        .excel-table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 0.9rem;
        }}
        .excel-cell {{
            padding: 8px 12px;
            border: 1px solid rgba(255, 255, 255, 0.08);
            white-space: normal;
            word-wrap: break-word;
            max-width: 300px;
        }}
        tr:hover .excel-cell {{
            background-color: rgba(255, 255, 255, 0.01) !important;
        }}
        @media print {{
            @page {{
                size: A4 landscape;
                margin: 12mm 15mm 12mm 15mm;
            }}
            body {{
                background-color: white;
                color: black;
                padding: 0;
                margin: 0;
            }}
            .header, .tabs, .btn-print {{
                display: none;
            }}
            .container {{
                background: none;
                box-shadow: none;
                border: none;
                overflow: visible;
                max-width: 100%;
            }}
            .tab-content {{
                display: block !important;
                page-break-after: always;
                padding: 0;
                margin-bottom: 30px;
            }}
            .tab-content::before {{
                content: attr(data-sheet-name);
                display: block;
                font-size: 1.6rem;
                font-weight: bold;
                margin-bottom: 12px;
                color: #2C3E50;
                border-bottom: 2px solid #4A90E2;
                padding-bottom: 6px;
            }}
            .table-responsive {{
                overflow: visible;
                border: none;
                background: none;
            }}
            .excel-table {{
                width: 100%;
                color: black;
                page-break-inside: auto;
            }}
            tr {{
                page-break-inside: avoid;
                page-break-after: auto;
            }}
            .excel-cell {{
                border: 1px solid #999 !important;
                color: black !important;
                background-color: transparent !important;
                padding: 6px 8px;
                font-size: 0.85rem;
            }}
        }}
    </style>
</head>
<body>
    <div class="header">
        <h1>AMR Bill of Materials & Calculations</h1>
        <button class="btn-print" onclick="window.print()">Print / Export all as PDF</button>
    </div>
    
    <div class="container">
        <div class="tabs">
            {tab_headers_html}
        </div>
        
        {tab_contents_html}
    </div>

    <script>
        function openTab(evt, tabId) {{
            var i, tabcontent, tablinks;
            tabcontent = document.getElementsByClassName("tab-content");
            for (i = 0; i < tabcontent.length; i++) {{
                tabcontent[i].classList.remove("active");
            }}
            tablinks = document.getElementsByClassName("tab-link");
            for (i = 0; i < tablinks.length; i++) {{
                tablinks[i].classList.remove("active");
            }}
            document.getElementById(tabId).classList.add("active");
            evt.currentTarget.classList.add("active");
        }}
    </script>
</body>
</html>
"""
    with open(html_out, "w", encoding="utf-8") as f:
        f.write(html_page)
    print(f"Excel HTML Viewer generated successfully: {html_out}")

if __name__ == "__main__":
    build_report_html(
        "/home/charan/.gemini/antigravity-ide/scratch/tusk-amr-analysis/design_report.md",
        "/home/charan/.gemini/antigravity-ide/scratch/tusk-amr-analysis/design_report.html",
        "AMR Design & Mechanism Report"
    )
    build_report_html(
        "/home/charan/.gemini/antigravity-ide/scratch/tusk-amr-analysis/mechanisms_report.md",
        "/home/charan/.gemini/antigravity-ide/scratch/tusk-amr-analysis/mechanisms_report.html",
        "Pallet AMR Mechanisms & Components Report"
    )
    build_report_html(
        "/home/charan/.gemini/antigravity-ide/scratch/tusk-amr-analysis/sequential_slide_report.md",
        "/home/charan/.gemini/antigravity-ide/scratch/tusk-amr-analysis/sequential_slide_report.html",
        "Sequential Latching Telescopic Slide Report"
    )
    build_report_html(
        "/home/charan/.gemini/antigravity-ide/scratch/tusk-amr-analysis/stepper_drive_guide.md",
        "/home/charan/.gemini/antigravity-ide/scratch/tusk-amr-analysis/stepper_drive_guide.html",
        "Stepper Motor Drive Guide for Telescopic Slides"
    )
    build_report_html(
        "/home/charan/.gemini/antigravity-ide/scratch/tusk-amr-analysis/tusk_e_series_cad_design_guide.md",
        "/home/charan/.gemini/antigravity-ide/scratch/tusk-amr-analysis/tusk_e_series_cad_design_guide.html",
        "Tusk Robots E-Series CAD Design & Dimensioning Guide"
    )
    excel_to_html_tabs()
